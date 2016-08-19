import feedparser
import sqlite3
from time import gmtime, strftime
from openpyxl import Workbook

#establishing connection to the database and creating a cursor
conn = sqlite3.connect('scalertsgoogle.db')
c = conn.cursor()

#The table schema 
def create_table():
    c.execute("CREATE TABLE IF NOT EXISTS feeds(title TEXT, summary TEXT, link TEXT, datestamp TEXT)")

#inserting values per row of the database based on the parameters provided by the user, which are obtained after parsing the feed
def data_entry(title,summ, link, date):
    c.execute("INSERT INTO feeds(title,summary,link, datestamp) VALUES (?, ?, ?, ?)",(title, summ, link, date))
    conn.commit()

#creating the table
create_table()

#The list of rss feed urls from which we are parsing
urllist = ['https://www.google.com/alerts/feeds/04151699818659974044/832217289615740120']
urllist.append('https://www.google.com/alerts/feeds/04151699818659974044/17448190345750142348')


#create an excel workbook
wb = Workbook()

# grab the active worksheet
ws = wb.active

#populate the header of the worksheet
ws.cell(row=1,column=1).value = "Title"
ws.cell(row=1,column=2).value = "Summary"
ws.cell(row=1,column=3).value = "Link"
ws.cell(row=1,column=4).value = "Datestamp"

# function to check if the post is already in the database
def post_in_db(title):
    c.execute("SELECT * FROM feeds WHERE EXISTS(SELECT * FROM feeds WHERE title =?)",[title])
    exist = c.fetchone()
    if exist is None:
        #a post with the title does not exist
        return 0
    else:
        #a post with the title exists
        return 1


#function to parse url feeds and feed to the database
def parse_feed(url):
    feed= feedparser.parse(url)
    #inserting posts from feed into the database and the excel file only if post is not present in the database
    row = 2;
    for post in feed.entries:
        if not post_in_db(post.title):
            printToFile(post.title, post.summary, post.link, post.published,row)
            data_entry(post.title, post.summary, post.link, post.published)
            row = row+1

#function to insert posts from the list of urls
def insertPost():
    for i in range(0,len(urllist)):
        parse_feed(urllist[i])
    
#function to read the database feeds and print recent posts
def printToFile(title,summary,link,published,i):
    ws.cell(row=i,column=1).value = title
    # Save the file
    wb.save("scalertsgoogle.xlsx")
    
    ws.cell(row=i,column=2).value = summary
    # Save the file
    wb.save("scalertsgoogle.xlsx")

    ws.cell(row=i,column=3).value = link
    # Save the file
    wb.save("scalertsgoogle.xlsx")

    ws.cell(row=i,column=4).value = published
    # Save the file
    wb.save("scalertsgoogle.xlsx")
    
#insert posts into the database
insertPost()

#closing the cursor and database connection
c.close()
conn.close()
